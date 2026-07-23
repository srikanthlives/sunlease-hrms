"""
Generates IDFC FIRST Bank's bulk payment upload format for salary disbursement, using the
ACTUAL bank template (resources/bank_payment_template.xlsx, as provided by the bank) - same
headers, same instructions row, same column order; only the data rows are replaced.

Employees are split into two files by designation, matching how the bank file is actually
used at this organization:
  - "Coach Captain" designation  -> BLKPAYCC<MM><YYYY>.xlsx
  - everyone else                -> BLKPAYST<MM><YYYY>.xlsx

The amount paid per employee is the outstanding BALANCE for that month (net pay minus any
part-payments already recorded via /payments), so re-generating the file after some partial
payments have already gone out only asks the bank to pay what's still owed.
"""
import io
import zipfile
from pathlib import Path
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models, schemas
from ..database import get_db
from ..deps import require_payroll_or_admin

router = APIRouter(prefix="/bank-payments", tags=["bank-payments"])

RESOURCES_DIR = Path(__file__).resolve().parent.parent / "resources"
BANK_TEMPLATE_PATH = RESOURCES_DIR / "bank_payment_template.xlsx"
BANK_DATA_START_ROW = 3  # row 1 = headers, row 2 = instructions, data starts row 3

MONTH_ABBR = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _build_rows(db: Session, month: int, year: int, payload: schemas.BankPaymentGenerateRequest):
    remarks = payload.remarks or f"Salary for {MONTH_ABBR[month]} - {year}"
    cc_label = payload.coach_captain_designation.strip().lower()
    generation_mode = payload.generation_mode.lower()  # "full_salary" or "pending_payments"

    cc_rows, st_rows, skipped = [], [], []

    payslips = db.query(models.Payslip).filter_by(month=month, year=year).all()
    for ps in payslips:
        emp = db.query(models.Employee).get(ps.employee_id)
        if not emp:
            continue

        # Skip employees who have already been fully paid (balance <= 0)
        # This applies to both generation modes - we only include what's still owed
        paid_so_far = sum(
            p.amount for p in db.query(models.SalaryPayment).filter_by(
                employee_id=emp.id, month=month, year=year
            ).all()
        )
        balance = round(ps.net_pay - paid_so_far, 2)

        # For "pending_payments" mode, skip employees who are already fully paid
        if generation_mode == "pending_payments" and balance <= 0:
            continue

        if not emp.bank_account:
            skipped.append(f"{emp.employee_code}: no bank account number on file - excluded")
            continue

        row = (
            f"{emp.first_name} {emp.last_name}".strip(),  # Beneficiary Name
            emp.bank_account,                              # Beneficiary Account Number
            emp.ifsc or "",                                # IFSC
            "NEFT",                                        # Transaction Type
            payload.debit_account_number,                  # Debit Account Number
            payload.transaction_date,                      # Transaction Date (DD/MM/YYYY)
            balance,                                        # Amount (outstanding balance)
            "INR",                                          # Currency
            emp.email or "",                                # Beneficiary Email ID
            remarks,                                        # Remarks
            emp.employee_code,                              # Custom Header - 1 (Optional 1: employee id)
            emp.designation or "",                          # Custom Header - 2 (Optional 2: designation)
            "", "", "",                                     # Custom Header - 3/4/5 (unused)
        )
        if (emp.designation or "").strip().lower() == cc_label:
            cc_rows.append(row)
        else:
            st_rows.append(row)

        if not emp.ifsc:
            skipped.append(f"{emp.employee_code}: no IFSC on file - row included but IFSC is blank (required for NEFT)")

    return cc_rows, st_rows, skipped


def _sheet_bytes(rows: list[tuple]) -> bytes:
    if not BANK_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="Bank template file is missing on the server (app/resources/bank_payment_template.xlsx).")

    wb = openpyxl.load_workbook(BANK_TEMPLATE_PATH)
    ws = wb["Sheet1"]

    # Clear any example/leftover data rows below the header + instructions rows.
    if ws.max_row >= BANK_DATA_START_ROW:
        ws.delete_rows(BANK_DATA_START_ROW, ws.max_row - BANK_DATA_START_ROW + 1)

    for i, row in enumerate(rows):
        row_num = BANK_DATA_START_ROW + i
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/preview")
def preview_bank_payments(payload: schemas.BankPaymentGenerateRequest, db: Session = Depends(get_db), _=Depends(require_payroll_or_admin)):
    """Summary counts/warnings before committing to a download - no file is generated here."""
    cc_rows, st_rows, skipped = _build_rows(db, payload.month, payload.year, payload)
    amount_index = 6  # position of "Amount" within each row tuple
    return {
        "generation_mode": payload.generation_mode,
        "coach_captain_count": len(cc_rows),
        "coach_captain_total": round(sum(r[amount_index] for r in cc_rows), 2),
        "staff_count": len(st_rows),
        "staff_total": round(sum(r[amount_index] for r in st_rows), 2),
        "warnings": skipped,
    }


@router.post("/download")
def download_bank_payments(payload: schemas.BankPaymentGenerateRequest, db: Session = Depends(get_db), _=Depends(require_payroll_or_admin)):
    """Returns a ZIP containing BLKPAYCC<MM><YYYY>.xlsx and/or BLKPAYST<MM><YYYY>.xlsx."""
    cc_rows, st_rows, skipped = _build_rows(db, payload.month, payload.year, payload)

    if not cc_rows and not st_rows:
        raise HTTPException(status_code=400, detail="No employees with an outstanding balance and a bank account were found for this period.")

    suffix = f"{payload.month:02d}{payload.year}"
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if cc_rows:
            zf.writestr(f"BLKPAYCC{suffix}.xlsx", _sheet_bytes(cc_rows))
        if st_rows:
            zf.writestr(f"BLKPAYST{suffix}.xlsx", _sheet_bytes(st_rows))
    zip_buf.seek(0)

    headers = {
        "Content-Disposition": f'attachment; filename="BankPayments_{suffix}.zip"',
        "X-CC-Count": str(len(cc_rows)),
        "X-ST-Count": str(len(st_rows)),
        "X-Warning-Count": str(len(skipped)),
    }
    if skipped:
        # Header values must be latin-1 safe; url-encode and cap length defensively.
        joined = " | ".join(skipped[:15])
        if len(skipped) > 15:
            joined += f" | (+{len(skipped) - 15} more)"
        headers["X-Warnings"] = quote(joined)

    return StreamingResponse(
        zip_buf, media_type="application/zip", headers=headers,
    )
