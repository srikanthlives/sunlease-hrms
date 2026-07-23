"""
Compliance: PF (EPFO), ESI (ESIC), and Mediclaim tracking + downloadable filing templates.

The ESI download populates the ACTUAL ESIC "Monthly Contribution" upload template
(resources/esi_template.xlsx, as provided) - same headers, same instructions/reason-code
reference sheet, same column order; only the data rows are replaced. Upload it to the ESIC
portal exactly as downloaded.

PF and Mediclaim don't have an official template on file yet, so those stay as practical
starting-point statements built from payroll data - verify column requirements against the
current EPFO/insurer specification before uploading those two.

ESI contribution amounts (when shown elsewhere, e.g. the CTC view) are read from payslip lines
with codes 'ESIC' / 'EMPLOYER_ESIC' - the ESIC template itself doesn't need these, since ESIC
calculates contributions from wages on their end.
"""
import io
import math
import zipfile
from pathlib import Path

import pandas as pd
import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import require_payroll_or_admin
from ..utils.eligibility import is_eligible_for_period, parse_year_month

router = APIRouter(prefix="/compliance", tags=["compliance"])

RESOURCES_DIR = Path(__file__).resolve().parent.parent / "resources"
ESI_TEMPLATE_PATH = RESOURCES_DIR / "esi_template.xlsx"
PF_TEMPLATE_PATH = RESOURCES_DIR / "Apr_PF.xlsx"
PF_TEXT_TEMPLATE_PATH = RESOURCES_DIR / "Apr_PF.txt"
ESI_DATA_START_ROW = 2  # row 1 is the header; data starts immediately, no instructions row


def _line_amount(payslip: models.Payslip, code: str) -> float | None:
    for line in payslip.lines:
        if line.component_code.upper() == code.upper():
            return line.amount
    return None


def _xlsx_response(buf: io.BytesIO, filename: str) -> StreamingResponse:
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _df_xlsx_response(df: pd.DataFrame, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    return _xlsx_response(buf, filename)


def _ddmmyyyy(date_str: str | None) -> str:
    if not date_str:
        return ""
    parts = date_str.split("-")
    if len(parts) != 3:
        return ""
    year, month, day = parts
    return f"{day}/{month}/{year}"


def _get_model_rows(db, model, *, month: int | None = None, year: int | None = None):
    attr_name = f"_{model.__name__.lower()}s"
    if hasattr(db, attr_name):
        rows = getattr(db, attr_name)
        if month is not None and year is not None:
            return [
                row for row in rows
                if not hasattr(row, "month") or not hasattr(row, "year") or (
                    getattr(row, "month", None) == month and getattr(row, "year", None) == year
                )
            ]
        return rows

    query = db.query(model)
    if month is not None and year is not None:
        query = query.filter_by(month=month, year=year)
    return query.all()


def _get_model_by_id(db, model, row_id: int):
    attr_name = f"_{model.__name__.lower()}s"
    if hasattr(db, attr_name):
        rows = getattr(db, attr_name)
        return next((row for row in rows if getattr(row, "id", None) == row_id), None)
    return db.query(model).get(row_id)


def _round_currency(value: float | int | None) -> float:
    return round(float(value or 0), 0)


def _build_pf_rows(db: Session, month: int, year: int) -> list[dict]:
    rows = []
    payslips = _get_model_rows(db, models.Payslip, month=month, year=year)
    for ps in payslips:
        emp = _get_model_by_id(db, models.Employee, ps.employee_id)
        if not emp or not is_eligible_for_period(emp, month, year):
            continue
        if not emp.pf_eligible:
            continue

        basic_da = 0.0
        for line in ps.lines:
            if line.component_code.upper() in {"BASIC", "DA"}:
                basic_da += line.amount or 0

        gross_wages = _round_currency(basic_da)
        epf_wages = _round_currency(min(15000.0, gross_wages))
        eps_wages = _round_currency(epf_wages) if emp.eps_eligible else 0.0
        edli_wages = _round_currency(epf_wages)
        epf_contri = _round_currency(epf_wages * 0.12)
        eps_contri = _round_currency(eps_wages * 0.0833)
        epf_eps_diff = _round_currency(epf_contri - eps_contri)

        rows.append({
            "OPTEDOUT": "Y" if not emp.eps_eligible else "N",
            "UAN": emp.uan or "",
            "MEMBER NAME": f"{emp.first_name} {emp.last_name}".strip(),
            "GROSS WAGES": gross_wages,
            "EPF WAGES": epf_wages,
            "EPS WAGES": eps_wages,
            "EDLI WAGES": edli_wages,
            "EPF CONTRI REMITTED": epf_contri,
            "EPS CONTRI REMITTED": eps_contri,
            "EPF EPS DIFF REMITTED": epf_eps_diff,
            "NCP DAYS": 0,
            "REFUND OF ADVANCES": 0,
        })
    return rows


def _build_pf_text_content(rows: list[dict]) -> str:
    lines = []
    for row in rows:
        values = [
            row.get("UAN", ""),
            row.get("MEMBER NAME", ""),
            row.get("GROSS WAGES", ""),
            row.get("EPF WAGES", ""),
            row.get("EPS WAGES", ""),
            row.get("EDLI WAGES", ""),
            row.get("EPF CONTRI REMITTED", ""),
            row.get("EPS CONTRI REMITTED", ""),
            row.get("EPF EPS DIFF REMITTED", ""),
            row.get("NCP DAYS", ""),
            row.get("REFUND OF ADVANCES", ""),
        ]
        lines.append("#~#".join(str(v) for v in values))
    return "\n".join(lines) + ("\n" if lines else "")


@router.get("/pf-template")
def download_pf_template(month: int, year: int, db: Session = Depends(get_db), _=Depends(require_payroll_or_admin)):
    rows = _build_pf_rows(db, month, year)

    if PF_TEMPLATE_PATH.exists():
        wb = openpyxl.load_workbook(PF_TEMPLATE_PATH)
        ws = wb.active
        if ws.max_row > 1:
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row, 1)

        for idx, row in enumerate(rows, start=2):
            ws.cell(row=idx, column=1, value=row.get("OPTEDOUT", ""))
            ws.cell(row=idx, column=2, value=row.get("UAN", ""))
            ws.cell(row=idx, column=3, value=row.get("MEMBER NAME", ""))
            ws.cell(row=idx, column=4, value=row.get("GROSS WAGES", ""))
            ws.cell(row=idx, column=5, value=row.get("EPF WAGES", ""))
            ws.cell(row=idx, column=6, value=row.get("EPS WAGES", ""))
            ws.cell(row=idx, column=7, value=row.get("EDLI WAGES", ""))
            ws.cell(row=idx, column=8, value=row.get("EPF CONTRI REMITTED", ""))
            ws.cell(row=idx, column=9, value=row.get("EPS CONTRI REMITTED", ""))
            ws.cell(row=idx, column=10, value=row.get("EPF EPS DIFF REMITTED", ""))
            ws.cell(row=idx, column=11, value=row.get("NCP DAYS", ""))
            ws.cell(row=idx, column=12, value=row.get("REFUND OF ADVANCES", ""))

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)

        text_content = _build_pf_text_content(rows).encode("utf-8")
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"PF_Contribution_Statement_{month:02d}{year}.xlsx", xlsx_buf.getvalue())
            zf.writestr(f"PF_Contribution_Statement_{month:02d}{year}.txt", text_content)
        zip_buf.seek(0)
        return StreamingResponse(
            zip_buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="PF_Contribution_Statement_{month:02d}{year}.zip"'},
        )

    df = pd.DataFrame(rows, columns=[
        "OPTEDOUT", "UAN", "MEMBER NAME", "GROSS WAGES", "EPF WAGES", "EPS WAGES", "EDLI WAGES",
        "EPF CONTRI REMITTED", "EPS CONTRI REMITTED", "EPF EPS DIFF REMITTED",
        "NCP DAYS", "REFUND OF ADVANCES",
    ])
    xlsx_buf = io.BytesIO()
    with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    text_content = _build_pf_text_content(rows).encode("utf-8")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"PF_Contribution_Statement_{month:02d}{year}.xlsx", xlsx_buf.getvalue())
        zf.writestr(f"PF_Contribution_Statement_{month:02d}{year}.txt", text_content)
    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="PF_Contribution_Statement_{month:02d}{year}.zip"'},
    )


@router.get("/esi-template")
def download_esi_template(month: int, year: int, db: Session = Depends(get_db), _=Depends(require_payroll_or_admin)):
    """
    Fills the real ESIC Monthly Contribution upload template with this month's data:
    IP Number, IP Name, No. of days wages paid/payable, Total Monthly Wages, Reason Code
    (0 = normal; 2 = "Left Service" with Last Working Day filled in, if applicable this month).
    """
    if not ESI_TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="ESI template file is missing on the server (app/resources/esi_template.xlsx).")

    wb = openpyxl.load_workbook(ESI_TEMPLATE_PATH)
    ws = wb["Sheet1"]

    # Clear any example/leftover data rows below the header before writing fresh data.
    if ws.max_row >= ESI_DATA_START_ROW:
        ws.delete_rows(ESI_DATA_START_ROW, ws.max_row - ESI_DATA_START_ROW + 1)

    payslips = db.query(models.Payslip).filter_by(month=month, year=year).all()
    skipped = []
    row_num = ESI_DATA_START_ROW
    for ps in payslips:
        emp = db.query(models.Employee).get(ps.employee_id)
        if not emp or not is_eligible_for_period(emp, month, year):
            continue
        if not emp.esi_number:
            skipped.append(emp.employee_code)
            continue

        days = min(math.ceil(ps.present_days), int(ps.total_days)) if ps.present_days else 0

        left_this_month = False
        dol = parse_year_month(emp.date_of_leaving)
        if dol and dol == (year, month):
            left_this_month = True

        ws.cell(row=row_num, column=1, value=emp.esi_number)
        ws.cell(row=row_num, column=2, value=f"{emp.first_name} {emp.last_name}".strip())
        ws.cell(row=row_num, column=3, value=days)
        ws.cell(row=row_num, column=4, value=ps.gross_earnings)
        ws.cell(row=row_num, column=5, value=2 if left_this_month else 0)
        ws.cell(row=row_num, column=6, value=_ddmmyyyy(emp.date_of_leaving) if left_this_month else "")
        row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    response = _xlsx_response(buf, f"ESIC_MC_{month:02d}{year}.xlsx")
    if skipped:
        from urllib.parse import quote
        response.headers["X-Skipped-No-ESI-Number"] = quote(", ".join(skipped[:20]))
    return response


@router.get("/mediclaim-template")
def download_mediclaim_template(month: int, year: int, db: Session = Depends(get_db), _=Depends(require_payroll_or_admin)):
    rows = []
    employees = db.query(models.Employee).all()
    for emp in employees:
        if not is_eligible_for_period(emp, month, year):
            continue
        rows.append({
            "Employee Code": emp.employee_code,
            "Employee Name": f"{emp.first_name} {emp.last_name}".strip(),
            "Mediclaim Policy No.": emp.mediclaim_policy_no or "",
            "Department": emp.department or "",
            "Designation": emp.designation or "",
            "Date of Joining": emp.date_of_joining or "",
            "Date of Leaving": emp.date_of_leaving or "",
            "Endorsement Type": "Deletion" if emp.date_of_leaving else "Addition/Continuation",
        })
    df = pd.DataFrame(rows, columns=[
        "Employee Code", "Employee Name", "Mediclaim Policy No.", "Department", "Designation",
        "Date of Joining", "Date of Leaving", "Endorsement Type",
    ])
    return _df_xlsx_response(df, f"Mediclaim_Statement_{month:02d}{year}.xlsx")
