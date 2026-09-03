"""Bulk employee upload via .xlsx (blueprint §7 - creation wizard, batch
entry point). Each row creates a Draft Employee + EmploymentEpisode (same
status a single "New Employee" click starts at) with as much of Personal
Information / Address / Employment Information / Organizational
Assignment filled in as the row provides - Statutory/Bank/Documents/
Dependents/Nominees/Driving Licence stay per-employee, filled in later via
the wizard, since they don't fit a flat spreadsheet row well (multi-value
or file-upload fields).
"""
import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.validators import validate_aadhaar, validate_pan
from app.models.enums import AddressType, AuditAction, EpisodeStatus
from app.models.models import (
    Employee, EmploymentEpisode, Address, OrgAssignment,
    CostCenter, Department, Project, EmployeeCategory, EmployeeType, Designation, WorkLocation, User,
)
from app.services import audit_service, employee_service

# (field_key, column_header). Order here is the order columns appear in
# the downloadable template. "*" in the header marks a required field.
COLUMNS = [
    ("employee_number", "Employee Number*"),
    ("first_name", "First Name*"),
    ("middle_name", "Middle Name"),
    ("last_name", "Last Name*"),
    ("father_husband_name", "Father's/Husband's Name"),
    ("gender", "Gender (MALE/FEMALE/OTHER)"),
    ("date_of_birth", "Date of Birth (YYYY-MM-DD)"),
    ("marital_status", "Marital Status (SINGLE/MARRIED/OTHER)"),
    ("educational_qualification", "Educational Qualification"),
    ("mobile_number", "Mobile Number"),
    ("alternate_mobile_number", "Alternate Mobile Number"),
    ("personal_email", "Personal Email"),
    ("official_email", "Official Email"),
    ("aadhaar", "Aadhaar"),
    ("pan", "PAN"),
    ("emergency_contact_name", "Emergency Contact Name"),
    ("emergency_contact_relationship", "Emergency Contact Relationship"),
    ("emergency_contact_mobile", "Emergency Contact Mobile"),
    ("previous_designation", "Previous Designation"),
    ("previous_company_name", "Previous Company Name"),
    ("previous_company_details", "Previous Company Details"),
    ("previous_date_of_joining", "Previous Date of Joining (YYYY-MM-DD)"),
    ("total_experience_years", "Total Experience (years)"),
    ("present_line1", "Present Address Line 1"),
    ("present_line2", "Present Address Line 2"),
    ("present_city", "Present City"),
    ("present_state", "Present State"),
    ("present_pincode", "Present Pincode"),
    ("present_country", "Present Country"),
    ("same_as_present", "Permanent Same As Present (YES/NO)"),
    ("permanent_line1", "Permanent Address Line 1"),
    ("permanent_line2", "Permanent Address Line 2"),
    ("permanent_city", "Permanent City"),
    ("permanent_state", "Permanent State"),
    ("permanent_pincode", "Permanent Pincode"),
    ("permanent_country", "Permanent Country"),
    ("employment_type", "Employment Type (must match Organization Setup)"),
    ("employee_category", "Employee Category (must match Organization Setup)"),
    ("designation", "Designation (must match Organization Setup)"),
    ("work_location", "Work Location (must match Organization Setup)"),
    ("shift_group", "Shift Group"),
    ("date_of_joining", "Date of Joining (YYYY-MM-DD)"),
    ("confirmation_date", "Confirmation Date (YYYY-MM-DD)"),
    ("cost_center", "Cost Center (must match Organization Setup)"),
    ("department", "Department (must match Organization Setup)"),
    ("project", "Project (must match Organization Setup, optional)"),
    ("assignment_effective_from", "Assignment Effective From (YYYY-MM-DD)"),
]

SAMPLE_ROW = {
    "employee_number": "EMP00200", "first_name": "Ravi", "middle_name": "", "last_name": "Kumar",
    "father_husband_name": "Suresh Kumar", "gender": "MALE", "date_of_birth": "1992-05-14",
    "marital_status": "MARRIED", "educational_qualification": "B.Com",
    "mobile_number": "9876543210", "alternate_mobile_number": "", "personal_email": "ravi.k@example.com",
    "official_email": "ravi.kumar@company.com", "aadhaar": "234123412341", "pan": "ABCDE1234F",
    "emergency_contact_name": "Suresh Kumar", "emergency_contact_relationship": "Father",
    "emergency_contact_mobile": "9876500000",
    "previous_designation": "Driver", "previous_company_name": "ABC Transport",
    "previous_company_details": "Local bus operator", "previous_date_of_joining": "2018-06-01",
    "total_experience_years": 5.5,
    "present_line1": "12 MG Road", "present_line2": "", "present_city": "Puducherry",
    "present_state": "Puducherry", "present_pincode": "605001", "present_country": "India",
    "same_as_present": "YES", "permanent_line1": "", "permanent_line2": "", "permanent_city": "",
    "permanent_state": "", "permanent_pincode": "", "permanent_country": "",
    "employment_type": "Permanent", "employee_category": "Driver", "designation": "Bus Driver",
    "work_location": "Puducherry Depot", "shift_group": "General Shift",
    "date_of_joining": "2026-01-01", "confirmation_date": "",
    "cost_center": "Puducherry", "department": "Operations", "project": "",
    "assignment_effective_from": "2026-01-01",
}


def build_template_workbook(db: Session) -> Workbook:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Employees"

    headers = [header for _, header in COLUMNS]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    sheet.append([SAMPLE_ROW.get(key, "") for key, _ in COLUMNS])
    for col_idx, _ in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = 26

    ref = wb.create_sheet("Reference Values")
    ref.append(["Use these exact names in the matching columns on the Employees sheet"])
    ref["A1"].font = Font(bold=True)

    def _dump(row_label, col_letter, names):
        ref[f"{col_letter}1"] = row_label
        ref[f"{col_letter}1"].font = Font(bold=True)
        for i, name in enumerate(names, start=2):
            ref[f"{col_letter}{i}"] = name
        ref.column_dimensions[col_letter].width = 26

    ref.delete_rows(1)
    _dump("Cost Centers", "A", [c.name for c in db.query(CostCenter).filter(CostCenter.is_active.is_(True)).all()])
    _dump("Departments", "B", [d.name for d in db.query(Department).filter(Department.is_active.is_(True)).all()])
    _dump("Projects", "C", [p.name for p in db.query(Project).filter(Project.is_active.is_(True)).all()])
    _dump("Employee Categories", "D", [c.name for c in db.query(EmployeeCategory).filter(EmployeeCategory.is_active.is_(True)).all()])
    _dump("Employment Types", "E", [t.name for t in db.query(EmployeeType).filter(EmployeeType.is_active.is_(True)).all()])
    _dump("Designations", "F", [d.name for d in db.query(Designation).filter(Designation.is_active.is_(True)).all()])
    _dump("Work Locations", "G", [w.name for w in db.query(WorkLocation).filter(WorkLocation.is_active.is_(True)).all()])

    return wb


def _cell_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _cell_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _lookup(db: Session, model, name: str | None):
    if not name:
        return None
    return db.query(model).filter(model.name == name, model.is_active.is_(True)).first()


def import_workbook(db: Session, file_bytes: bytes, actor: User) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb["Employees"] if "Employees" in wb.sheetnames else wb.active

    header_row = [(_cell_str(c.value) or "") for c in sheet[1]]
    key_by_column_index = {}
    for idx, header in enumerate(header_row):
        for key, expected_header in COLUMNS:
            if header.strip().lower() == expected_header.strip().lower():
                key_by_column_index[idx] = key
                break

    created = 0
    errors = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue  # skip blank rows

        data = {}
        for idx, value in enumerate(row):
            key = key_by_column_index.get(idx)
            if key:
                data[key] = value

        employee_number = _cell_str(data.get("employee_number"))
        first_name = _cell_str(data.get("first_name"))
        last_name = _cell_str(data.get("last_name"))

        if not employee_number or not first_name or not last_name:
            errors.append({"row": row_number, "message": "Employee Number, First Name and Last Name are required"})
            continue

        savepoint = db.begin_nested()
        try:
            if db.query(EmploymentEpisode).filter(EmploymentEpisode.employee_number == employee_number).first():
                raise ValueError(f"Employee Number '{employee_number}' already in use")

            aadhaar = _cell_str(data.get("aadhaar"))
            if aadhaar:
                aadhaar = validate_aadhaar(aadhaar)
            pan = _cell_str(data.get("pan"))
            if pan:
                pan = validate_pan(pan)

            employee = Employee(
                first_name=first_name, middle_name=_cell_str(data.get("middle_name")), last_name=last_name,
                father_husband_name=_cell_str(data.get("father_husband_name")),
                gender=_cell_str(data.get("gender")), date_of_birth=_cell_date(data.get("date_of_birth")),
                marital_status=_cell_str(data.get("marital_status")),
                educational_qualification=_cell_str(data.get("educational_qualification")),
                mobile_number=_cell_str(data.get("mobile_number")),
                alternate_mobile_number=_cell_str(data.get("alternate_mobile_number")),
                personal_email=_cell_str(data.get("personal_email")), official_email=_cell_str(data.get("official_email")),
                aadhaar=aadhaar, pan=pan,
                emergency_contact_name=_cell_str(data.get("emergency_contact_name")),
                emergency_contact_relationship=_cell_str(data.get("emergency_contact_relationship")),
                emergency_contact_mobile=_cell_str(data.get("emergency_contact_mobile")),
                previous_designation=_cell_str(data.get("previous_designation")),
                previous_company_name=_cell_str(data.get("previous_company_name")),
                previous_company_details=_cell_str(data.get("previous_company_details")),
                previous_date_of_joining=_cell_date(data.get("previous_date_of_joining")),
                total_experience_years=_cell_float(data.get("total_experience_years")),
            )
            db.add(employee)
            db.flush()

            present_fields = {f: _cell_str(data.get(f"present_{f}")) for f in ("line1", "line2", "city", "state", "pincode", "country")}
            if any(present_fields.values()):
                db.add(Address(employee_id=employee.id, address_type=AddressType.PRESENT, **present_fields))

            same_as_present = (_cell_str(data.get("same_as_present")) or "").upper() in ("YES", "Y", "TRUE", "1")
            if same_as_present and any(present_fields.values()):
                db.add(Address(employee_id=employee.id, address_type=AddressType.PERMANENT, **present_fields))
            else:
                permanent_fields = {f: _cell_str(data.get(f"permanent_{f}")) for f in ("line1", "line2", "city", "state", "pincode", "country")}
                if any(permanent_fields.values()):
                    db.add(Address(employee_id=employee.id, address_type=AddressType.PERMANENT, **permanent_fields))

            employment_type = _lookup(db, EmployeeType, _cell_str(data.get("employment_type")))
            employee_category = _lookup(db, EmployeeCategory, _cell_str(data.get("employee_category")))
            designation = _lookup(db, Designation, _cell_str(data.get("designation")))
            work_location = _lookup(db, WorkLocation, _cell_str(data.get("work_location")))

            episode = EmploymentEpisode(
                employee_id=employee.id, employee_number=employee_number, status=EpisodeStatus.DRAFT,
                employment_type_id=employment_type.id if employment_type else None,
                employee_category_id=employee_category.id if employee_category else None,
                designation_id=designation.id if designation else None,
                work_location_id=work_location.id if work_location else None,
                shift_group=_cell_str(data.get("shift_group")),
                date_of_joining=_cell_date(data.get("date_of_joining")),
                confirmation_date=_cell_date(data.get("confirmation_date")),
            )
            db.add(episode)
            db.flush()

            cost_center = _lookup(db, CostCenter, _cell_str(data.get("cost_center")))
            department = _lookup(db, Department, _cell_str(data.get("department")))
            project = _lookup(db, Project, _cell_str(data.get("project")))
            if cost_center and department:
                employee_service.add_org_assignment(db, episode.id, {
                    "cost_center_id": cost_center.id, "department_id": department.id,
                    "project_id": project.id if project else None,
                    "effective_from": _cell_date(data.get("assignment_effective_from")) or date.today(),
                })

            audit_service.record(db, "EMPLOYEE_DRAFT", episode.id, AuditAction.CREATE, actor, new_value="bulk upload")
            savepoint.commit()
            created += 1
        except (ValueError, IntegrityError) as exc:
            savepoint.rollback()
            errors.append({"row": row_number, "message": str(exc.__cause__ or exc) if isinstance(exc, IntegrityError) else str(exc)})

    return {"created": created, "errors": errors}
